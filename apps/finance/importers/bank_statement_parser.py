import csv
from openpyxl import load_workbook
from datetime import datetime, date
from decimal import Decimal, InvalidOperation


class BankStatementParser:

    DATE_COLUMNS = {
        "date",
        "transaction_date",
        "txn_date",
        "transaction date",
        "txn date",
        "value_date",
        "value date",
    }

    DESCRIPTION_COLUMNS = {
        "description",
        "narration",
        "particulars",
        "remarks",
        "details",
        "transaction details",
    }

    DEBIT_COLUMNS = {
        "debit",
        "debit_amount",
        "debit amount",
        "withdrawal",
        "withdrawal_amount",
        "withdrawal amount",
        "withdrawal amt",
    }

    CREDIT_COLUMNS = {
        "credit",
        "credit_amount",
        "credit amount",
        "deposit",
        "deposit_amount",
        "deposit amount",
        "deposit amt",
    }

    BALANCE_COLUMNS = {
        "balance",
        "running_balance",
        "running balance",
        "closing_balance",
        "closing balance",
    }

    REFERENCE_COLUMNS = {
        "reference",
        "reference_number",
        "reference number",
        "ref no",
        "ref_no",
        "transaction id",
        "utr",
        "utr number",
        "cheque number",
    }

    @staticmethod
    def _normalize_header(
        value,
    ):
        if value is None:
            return ""

        return (
            str(value)
            .replace("\ufeff", "")
            .strip()
            .lower()
            .replace("-", " ")
            .replace("_", " ")
            .replace(".", "")
            .replace(":", "")
            .strip()
        )

    @staticmethod
    def _clean_text(
        value,
    ):
        if value is None:
            return ""

        return str(
            value
        ).strip()

    @staticmethod
    def _parse_decimal(
        value,
        field_name,
        default="0",
    ):
        if value is None:
            return Decimal(
                default
            )

        text = str(
            value
        ).strip()

        if not text:
            return Decimal(
                default
            )

        text = (
            text
            .replace(",", "")
            .replace("₹", "")
            .strip()
        )

        try:
            return Decimal(
                text
            )

        except (
            InvalidOperation,
            ValueError,
        ):
            raise ValueError(
                f"Invalid {field_name}: "
                f"{value}"
            )

    @staticmethod
    def _parse_date(
        value,
        field_name="transaction date",
    ):
        if value is None:
            raise ValueError(
                f"{field_name} is required."
            )

        if isinstance(
            value,
            datetime,
        ):
            return value

        if isinstance(
            value,
            date,
        ):
            return datetime.combine(
                value,
                datetime.min.time(),
            )

        text = str(
            value
        ).strip()

        if not text:
            raise ValueError(
                f"{field_name} is required."
            )

        formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%b-%Y",
            "%d %b %Y",
            "%d-%m-%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
        ]

        for date_format in formats:
            try:
                return datetime.strptime(
                    text,
                    date_format,
                )

            except ValueError:
                continue

        raise ValueError(
            f"Invalid {field_name}: "
            f"{value}"
        )

    @staticmethod
    def _find_column(
        headers,
        candidates,
    ):
        normalized = {
            BankStatementParser
            ._normalize_header(
                header
            ):
                header
            for header
            in headers
        }

        for candidate in candidates:
            candidate_normalized = (
                BankStatementParser
                ._normalize_header(
                    candidate
                )
            )

            if (
                candidate_normalized
                in normalized
            ):
                return normalized[
                    candidate_normalized
                ]

        return None

    @staticmethod
    def _detect_columns(
        headers,
    ):
        if not headers:
            raise ValueError(
                "Statement file has "
                "no headers."
            )

        date_column = (
            BankStatementParser
            ._find_column(
                headers,
                BankStatementParser
                .DATE_COLUMNS,
            )
        )

        description_column = (
            BankStatementParser
            ._find_column(
                headers,
                BankStatementParser
                .DESCRIPTION_COLUMNS,
            )
        )

        debit_column = (
            BankStatementParser
            ._find_column(
                headers,
                BankStatementParser
                .DEBIT_COLUMNS,
            )
        )

        credit_column = (
            BankStatementParser
            ._find_column(
                headers,
                BankStatementParser
                .CREDIT_COLUMNS,
            )
        )

        balance_column = (
            BankStatementParser
            ._find_column(
                headers,
                BankStatementParser
                .BALANCE_COLUMNS,
            )
        )

        reference_column = (
            BankStatementParser
            ._find_column(
                headers,
                BankStatementParser
                .REFERENCE_COLUMNS,
            )
        )

        if not date_column:
            raise ValueError(
                "Transaction date column "
                "not found."
            )

        if (
            not debit_column
            and not credit_column
        ):
            raise ValueError(
                "Debit/Credit columns "
                "not found."
            )

        return {
            "date":
                date_column,
            "description":
                description_column,
            "debit":
                debit_column,
            "credit":
                credit_column,
            "balance":
                balance_column,
            "reference":
                reference_column,
        }

    @staticmethod
    def _normalize_row(
        *,
        row,
        columns,
        line_number,
    ):
        transaction_date = (
            BankStatementParser
            ._parse_date(
                row.get(
                    columns["date"]
                )
            )
        )

        description = ""

        if columns[
            "description"
        ]:
            description = (
                BankStatementParser
                ._clean_text(
                    row.get(
                        columns[
                            "description"
                        ]
                    )
                )
            )

        external_reference = ""

        if columns[
            "reference"
        ]:
            external_reference = (
                BankStatementParser
                ._clean_text(
                    row.get(
                        columns[
                            "reference"
                        ]
                    )
                )
            )

        debit_amount = Decimal("0")

        if columns["debit"]:
            debit_amount = (
                BankStatementParser
                ._parse_decimal(
                    row.get(
                        columns[
                            "debit"
                        ]
                    ),
                    "debit amount",
                )
            )

        credit_amount = Decimal("0")

        if columns["credit"]:
            credit_amount = (
                BankStatementParser
                ._parse_decimal(
                    row.get(
                        columns[
                            "credit"
                        ]
                    ),
                    "credit amount",
                )
            )

        if (
            debit_amount < 0
            or credit_amount < 0
        ):
            raise ValueError(
                "Debit and credit amounts "
                "cannot be negative."
            )

        if (
            debit_amount > 0
            and credit_amount > 0
        ):
            raise ValueError(
                "Statement line cannot "
                "contain both debit "
                "and credit."
            )

        if (
            debit_amount == 0
            and credit_amount == 0
        ):
            raise ValueError(
                "Statement line must "
                "contain debit or credit."
            )

        running_balance = None

        if columns["balance"]:
            value = row.get(
                columns[
                    "balance"
                ]
            )

            if (
                value is not None
                and str(value).strip()
            ):
                running_balance = (
                    BankStatementParser
                    ._parse_decimal(
                        value,
                        "running balance",
                    )
                )

        return {
            "line_number":
                str(line_number),
            "transaction_date":
                transaction_date,
            "value_date":
                None,
            "description":
                description,
            "external_reference":
                external_reference,
            "debit_amount":
                debit_amount,
            "credit_amount":
                credit_amount,
            "running_balance":
                running_balance,
        }

    @staticmethod
    def parse_csv(
        file_path,
    ):
        rows = []

        with open(
            file_path,
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as file:
            reader = csv.DictReader(
                file
            )

            if not reader.fieldnames:
                raise ValueError(
                    "CSV file has no headers."
                )

            columns = (
                BankStatementParser
                ._detect_columns(
                    reader.fieldnames
                )
            )

            line_number = 0

            for raw_row in reader:
                if not any(
                    str(value).strip()
                    for value
                    in raw_row.values()
                    if value is not None
                ):
                    continue

                line_number += 1

                normalized = (
                    BankStatementParser
                    ._normalize_row(
                        row=raw_row,
                        columns=columns,
                        line_number=(
                            line_number
                        ),
                    )
                )

                rows.append(
                    normalized
                )

        if not rows:
            raise ValueError(
                "CSV statement contains "
                "no transaction rows."
            )

        return rows

    @staticmethod
    def parse_xlsx(
        file_path,
    ):
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

        rows_iterator = (
            worksheet.iter_rows(
                values_only=True
            )
        )

        try:
            header_row = next(
                rows_iterator
            )

        except StopIteration:
            raise ValueError(
                "XLSX file is empty."
            )

        headers = [
            (
                str(value).strip()
                if value is not None
                else ""
            )
            for value in header_row
        ]

        if not any(
            headers
        ):
            raise ValueError(
                "XLSX file has no headers."
            )

        columns = (
            BankStatementParser
            ._detect_columns(
                headers
            )
        )

        rows = []

        line_number = 0

        for values in rows_iterator:
            if not any(
                value is not None
                and str(value).strip()
                for value in values
            ):
                continue

            row = {
                headers[index]:
                    values[index]
                    if index < len(values)
                    else None
                for index
                in range(
                    len(headers)
                )
            }

            line_number += 1

            normalized = (
                BankStatementParser
                ._normalize_row(
                    row=row,
                    columns=columns,
                    line_number=(
                        line_number
                    ),
                )
            )

            rows.append(
                normalized
            )

        workbook.close()

        if not rows:
            raise ValueError(
                "XLSX statement contains "
                "no transaction rows."
            )

        return rows